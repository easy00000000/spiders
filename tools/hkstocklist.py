# -*- coding: utf-8 -*-
"""
Created on Tue Apr 17 13:55:15 2018

@author: Easy
"""
from urllib.request import urlopen
from data import read_stocklist

def parse_stocklist(stocklist_url = 'http://www.hkex.com.hk/eng/services/trading/securities/securitieslists/',
                    stocklist_filename = 'ListOfSecurities.xlsx'):
    url = stocklist_url + stocklist_filename
    with urlopen(url) as response, open(stocklist_filename, 'wb') as out_file:
        data = response.read()
        out_file.write(data)
        
def xlsx2json(xlsx_file = 'ListOfSecurities.xlsx',
              json_file = 'stocklist.json',
              categories = ('Equity','Real Estate Investment Trusts')):
    import json
    from openpyxl import load_workbook
    
    wb = load_workbook(xlsx_file)
    ws = wb['ListOfSecurities']
    
    stocklist = []
    row = 4
    col = 1    
    
    while (ws.cell(row,col).value is not None):
        if (ws.cell(row,col+2).value in tuple(categories)):
            stock = {'Code': ws.cell(row,col).value,
                     'Name': ws.cell(row,col+1).value,
                     'Category': ws.cell(row,col+2).value, # Equity or Real Estate Investment Trust
                     'ISIN': ws.cell(row,col+6).value}
            stocklist.append(stock)
        row = row + 1
        
    with open(json_file, 'w') as outfile:  
        json.dump(stocklist, outfile)

def main():
    print('Load updated stocklist from HKEx.')
    parse_stocklist()
    print('Save to Json file.')
    xlsx2json()
    #print('Read Stock Code from Json file')
    #s = read_stocklist()
    #print(s)
        
if __name__ == "__main__":
    main()